import openpyxl
import netaddr


def open_excel_routers(file):
    wb = openpyxl.load_workbook(file)
    for sheet_name in wb.sheetnames:
        addresses = []
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cells in row:
                addresses.append(cells.value)
        yield addresses

def get_excel_sheets(file):
    wb = openpyxl.load_workbook(file)
    return wb.sheetnames

def save_output_to_excel(output_list, file_name):
    from openpyxl import Workbook
    wb = Workbook()
    for position, sheet_value in enumerate(output_list):
        sheet = wb.create_sheet(sheet_value, position)
        for position, output_list_value in enumerate(output_list[sheet_value]):
            sheet.cell(row=position+1, column=1).value = output_list[sheet_value][position]
    wb.save(file_name)




file_name = 'output_aggr.xlsx'
opcos_address = open_excel_routers('список публичных префиксов.xlsx')
sheet_name = get_excel_sheets('список публичных префиксов.xlsx')
list_network_full = dict.fromkeys(sheet_name)
for position, address in enumerate(opcos_address):
    list_netw_aggr = netaddr.cidr_merge(address)
    list_network = [str(network) for network in list_netw_aggr]
    list_network_full[sheet_name[position]] = list_network
save_output_to_excel(list_network_full, file_name)
