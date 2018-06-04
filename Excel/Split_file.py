import openpyxl

def open_excel_routers(file):
    addresses = []
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    for row in sheet.iter_rows():
        addresses.append(row)
    return addresses


def save_output_to_excel(output_list, file_name):
    from openpyxl import Workbook
    wb = Workbook()
    sheet = wb.active
    for i in range(0, len(output_list)):
        for k in range(0, len(output_list[i])):
            sheet.cell(row=output_list.index(output_list[i])+1, column=(k+1)).value = output_list[i][k]
    wb.save(file_name)

f2_value = []
line_1_value = []
file_name = 'output_split.xlsx'
file1 = 'output_Minsk.xlsx'
file2 = 'output_opt_Minsk.xlsx'
f1 = open_excel_routers(file1)
f2 = open_excel_routers(file2)
for line_1 in f1:
    for line_2 in f2:
        if line_1[0].value == line_2[0].value and line_1[5].value == line_2[3].value:
            line_1_value = []
            for cells in line_1:
                line_1_value.append(cells.value)
            line_1_value.append(line_2[1].value)
            line_1_value.append(line_2[2].value)
            f2_value.append(line_1_value)
save_output_to_excel(f2_value, file_name)
