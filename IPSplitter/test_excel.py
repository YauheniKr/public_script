# -*- coding: utf-8 -*-
import openpyxl
import sys
from IPSplitter import IPSplitter

sites = []
hostnumber = []
host_bin = []
host_bin_less = []
subnet = []
subnet_del = []
filename = sys.argv[1]
wb = openpyxl.load_workbook(filename)
sheet = wb.get_sheet_by_name('Лист1')
ipv4_network = IPSplitter(sheet['C1'].value)


for row in range(3, sheet.max_row + 1):
    sites.append(sheet['a' + str(row)].value)
    hostnumber.append(sheet['b' + str(row)].value)
    host_value = sheet['b' + str(row)].value
    host_bin_value = len(bin(host_value)[2::])
    if 2**host_bin_value - 2 < host_value:
        host_bin_value = host_bin_value + 1
    host_bin.append(host_bin_value)

host_bin_tup = tuple(host_bin)
host_bin = sorted(host_bin, reverse=True)
sites_netw = dict.fromkeys(sites)

for host in host_bin:
    if host < 3:
        subnet.extend(ipv4_network.get_subnet(29, count=1))
    else:
        subnet.extend(ipv4_network.get_subnet(32 - host, count=1))

for hosts in host_bin_tup:
    i = 0
    for net in subnet:
        if hosts == (32 - net.prefixlen) and i == 0:
            subnet_del.append(net)
            subnet.remove(net)
        elif hosts < 3 and (net.prefixlen) == 29 and i == 0:
            subnet_del.append(net)
            subnet.remove(net)
        elif hosts != (32 - net.prefixlen) and i == 0:
            i = 0
            continue
        i += 1
i = 0
for row in range(3, sheet.max_row + 1):
    sheet['C' + str(row)].value = str(subnet_del[i])
    i += 1
wb.save(filename)
print('Разбивка сети на подсети завершена')


















