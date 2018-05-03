import openpyxl
import sys
import getpass
import telnetlib
import time
import re
import socket
from datetime import datetime
from tabulate import tabulate
import threading


def open_excel_routers(file):
    addresses = []
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_active_sheet()
    for row in range(1, sheet.max_row+1):
        addresses.append(sheet['A'+str(row)].value)
    return addresses

def save_output_to_excel(output_list, file_name):
    from openpyxl import Workbook
    wb = Workbook()
    sheet = wb.active
    for i in range(0, len(output_list)):
        for k in range(0, 4):
            sheet.cell(row=output_list.index(output_list[i])+1, column=(k+1)).value = output_list[i][k]
    wb.save(file_name)


def parser_show_interface_description(output, template):
    import textfsm
    result = []
    f = open(template)
    re_table = textfsm.TextFSM(f)
    result_header = re_table.header
    result.append(result_header)
    result_output = re_table.ParseText(output)
    result.extend(result_output)
    return result


def conn_threads(function, devices, command, Username, Password, limit=5):
    from queue import Queue


    devices_groups = (devices[idx:idx+limit]
                      for idx in range(0, len(devices), limit))

    for group in devices_groups:
        print(group)
        threads = []
        q = Queue()

        for device in devices:
            # Передаем очередь как аргумент, функции
            th = threading.Thread(target=function, args=(device, command, Username, Password, q))
            th.start()
            threads.append(th)

        for th in threads:
            th.join()

    results = []
    # Берем результаты из очереди и добавляем их в список results
    for t in threads:
        results.append(q.get())

    return results

"""
def send_show_command(address_list, command, Username, Password):
    output_list = []
    for IP in address_list:
        print('Connection to device {}'.format(IP))
        with telnetlib.Telnet(IP) as t:
            t.read_until(b'Username:')
            t.write(Username + b'\n')

            t.read_until(b'Password:')
            t.write(Password + b'\n')
            #t.write(b'enable\n')

            #t.read_until(b'Password:')
            #t.write(ENABLE_PASS + b'\n')
            t.write(b'terminal length 0\n')
            t.write(command + b'\n')

            time.sleep(20)

            output = t.read_very_eager().decode('utf-8')
            output_list.append(output.strip())
    return output_list
"""


def send_show_command(IP, command, Username, Password, queue):
    output_list = []
    #for IP in address_list:
    print('Connection to device {}'.format(IP))
    with telnetlib.Telnet(IP, timeout=5) as t:
        t.read_until(b'Username:')
        t.write(Username + b'\n')

        t.read_until(b'Password:')
        t.write(Password + b'\n')
        t.write(b'_cmdline-mode on\n')

        #t.read_until(b'Password:')
        t.write(b'Y\n')
        t.read_until(b'Please input password:')
        #t.write(b'Jinhua1920unauthorized\n')
        t.write(b'512900\n')
        t.write(b'screen-length disable\n')

        time.sleep(1)

        #t.read_very_eager()
        t.write(command + b'\n')
        output = ''
        while True:
            try:
                part = t.read_some().decode('utf-8')
            except socket.timeout:
                break
            output += part
        #output = t.read_very_eager().decode('utf-8')
        #output_list.append(output)
        queue.put({IP: output})


    #return output_list


device_list = []
device = []
device_template = ['device_type', 'ip', 'username', 'password', 'enable']
command = 'display interface brief'.encode('utf-8')

print('Insert Username and Password')
Username = input('Username:').encode('utf-8')
Password = getpass.getpass().encode('utf-8')
#print('Insert command for send to routers')
#command = input('Command:').encode('utf-8')
#enable_Password = Password
#device_type = input('Device Type(cisco_ios for ssh or cisco_ios_telnet for telnet ):')
#device.append(device_type)

output_file_name = 'output_cor_full_1.xlsx'
template = 'template.txt'
file = sys.argv[1]
address_list = open_excel_routers(file)

start_time = datetime.now()
#output_command = send_show_command(address_list, command, Username, Password)
output_command = conn_threads(send_show_command, address_list, command, Username, Password)
#output_command_cleared = re.sub('(\r\n {66})', '', ''.join(output_command))
listing_out = []


for dict in output_command:
    #print('\r\n'.join(list(dict.values())))
    listing = parser_show_interface_description('\r\n'.join(list(dict.values())), template)
    listing_out.extend(listing)
print(tabulate(listing_out[1:], headers=listing_out[0]))
output_file = save_output_to_excel(listing_out, output_file_name)

print(datetime.now() - start_time)
