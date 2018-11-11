#-*- coding: utf-8 -*-

import netmiko
from netmiko import ConnectHandler
from datetime import datetime
import getpass
import openpyxl
import re
import netmiko.ssh_exception
import clitable
from concurrent.futures import ThreadPoolExecutor, as_completed
import sys
import logbook

app_log = logbook.Logger('App')

class RouterSSH:
    def __init__(self, **device_dict):
        self.device_dict = device_dict
        self.ssh = ConnectHandler(**device_dict)

    def send_command(self, command):
        app_log.trace('Send command {} to device {} \n'.format(command, self.device_dict['ip']))
        if not self.ssh.check_enable_mode:
            self.ssh.enable()
        '''
        if self.device_dict['device_type'] == 'hp':
            self.ssh.send_command('_cmdline-mode on \n Y\nJinhua1920unauthorized')
        '''
        return self.ssh.send_command(command, strip_prompt=False)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, traceback):
        print('Закрываю соединение с устройством: {} \n'.format(self.device_dict['ip']))
        app_log.trace('Закрываю соединение с устройством: {} \n'.format(self.device_dict['ip']))
        self.ssh.disconnect()


def connect_ssh(device_dict, command, ):
    app_log.trace('Connection to device: {}'.format(device_dict['ip']))
    print('Connection to device: {}'.format(device_dict['ip']))
    with RouterSSH(**device_dict) as session:
        result = session.send_command(command)
    return {device_dict['ip']: result}


def parser_show_interface_description_clitable (output, command):
    result = []
    cli_table = clitable.CliTable('index', 'templates')
    attributes = {'Command': command, 'Vendor': 'Cisco'}
    cli_table.ParseCmd(output, attributes)
    data_rows = [list(row) for row in cli_table]
    header = list(cli_table.header)
    result.append(header)
    result.extend(list(data_rows))
    return result


def open_excel_file(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    for column in sheet.rows:
        yield [address.value for address in column]


def save_output_to_file(output_list, file_name):
    from openpyxl import Workbook
    wb = Workbook()
    sheet = wb.active
    for i in range(0, len(output_list)):
        for k in range(0, len(output_list[i])):
            sheet.cell(row=(i+1), column=(k+1)).value = output_list[i][k]
    wb.save(file_name)


def threads_conn(function, devices, command, limit=2):
    all_results = {}
    with ThreadPoolExecutor(max_workers=limit) as executor:
        future_ssh = [executor.submit(function, device, command) for device in devices]
        for f in as_completed(future_ssh):
            try:
                result = f.result()
            except netmiko.ssh_exception.NetMikoAuthenticationException as e:
                app_log.trace(e)
                print(e)
                break
            except netmiko.ssh_exception.NetMikoTimeoutException as e:
                app_log.trace(e)
                print(e)
            else:
                all_results.update(result)
        for f in reversed(future_ssh):
            if not f.cancel():
                try:
                    all_results.update(f.result())
                except netmiko.ssh_exception.SSHException:
                    pass
    return all_results


def router_params(routers_parameter):
    for router in routers_parameter:
        router.extend(device_values_kn)
        device_dict = {k: v for (k, v) in zip(device_template, router)}
        yield device_dict


def init_logging(filename: str = None):
    level = logbook.TRACE

    if filename:
        logbook.TimedRotatingFileHandler(filename, level=level).push_application()
    else:
        logbook.StreamHandler(sys.stdout, level=level).push_application()

    msg = 'Logging initialized, level: {}, mode: {}'.format(
        level,
        "stdout mode" if not filename else 'file mode: ' + filename
    )
    logger = logbook.Logger('Startup')
    logger.notice(msg)


device_template = ['ip', 'device_type', 'username', 'password', 'secret']

start_msg = '===> {} Connection to device: {}'
received_msg = '<=== {} Received result from device: {}'


Username = input('Username:')
Password = getpass.getpass()
print('Insert command for send to routers')
command = input('Command:')
device_values_kn = [Username, Password, Password]
#output_file_name = 'output_Minsk.xlsx'
file = 'routers_1.xlsx'
init_logging('movie-app.log')

routers_parameter = open_excel_file(file)
router = router_params(routers_parameter)
all_done = threads_conn(connect_ssh, router, command)
print(all_done)

'''
listing = parser_show_interface_description_clitable(re.sub('(\r\n {66})*', '', ''.join(list(all_done.values()))), command)
listing_out.extend(listing)

print(listing_out)

#output_file = save_output_to_file(listing_out, output_file_name)
'''