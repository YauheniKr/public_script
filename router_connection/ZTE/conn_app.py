import openpyxl
import clitable
import getpass


def open_excel_file(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    for row in sheet.rows:
        yield [address.value for address in row]


def save_output_to_file(output_list, file_name):
    from openpyxl import Workbook
    wb = Workbook()
    sheet = wb.active
    for i in range(0, len(output_list)):
        for k in range(0, len(output_list[i])):
            sheet.cell(row=(i+1), column=(k+1)).value = output_list[i][k]
    wb.save(file_name)


def router_params(routers_parameter):
    Username, Password = input_credentials()
    device_values_kn = [Username, Password, Password]
    device_template = ['ip', 'device_type', 'command', 'username', 'password', 'secret']
    for router in routers_parameter:
        router.extend(device_values_kn)
        device_dict = {k: v for (k, v) in zip(device_template, router)}
        yield device_dict


def parser_show_clitable(output):
    cli_table = clitable.CliTable('index', 'templates')
    for router in output:
        result = []
        attributes = {'Command': router['command'], 'Vendor': 'Cisco'}
        cli_table.ParseCmd(router['output'], attributes)
        data_rows = [list(row) for row in cli_table]
        header = list(cli_table.header)
        result.append(header)
        result.extend(list(data_rows))
        yield result






def input_credentials():
    Username = input('Username:')
    Password = getpass.getpass()
    return Username, Password

def take_command_line():
    print('Insert command for send to routers')
    command = input('Command:')
    return command

#TODO: Сделать распознавание команды из файла

#TODO: Сделать парсинг конфига с распознаванием выполненой команды из передаваемого словаря

#TODO: Сделать определение вендора для парсинга по файлу(или словарю) справочнику

#TODO: Оптимизировать функцию записи информации с маршрутизаторов в файл

