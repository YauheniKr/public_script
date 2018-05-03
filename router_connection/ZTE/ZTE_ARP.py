import openpyxl
import sys
import getpass
import telnetlib
import time
import re
import threading
from datetime import datetime
from tabulate import tabulate


def open_excel_routers(file):
    addresses = []
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    for row in range(1, sheet.max_row+1):
        addresses.append(sheet['A'+str(row)].value)
    return addresses

def save_output_to_excel(output_list, file_name):
    from openpyxl import Workbook
    wb = Workbook()
    sheet = wb.active
    for i in range(0, len(output_list)):
        for k in range(0, 7):
            sheet.cell(row=(i+1), column=(k+1)).value = output_list[i][k]
    wb.save(file_name)

"""
def device_dict(device_list, device_template):
    device_dict_list = []
    for dev in device_list:
        device_dic = dict.fromkeys(device_template)
        for key in list(device_dic.keys()):
            device_dic[key] = dev[(device_template.index(key))]
            device_dict_list.append(device_dic)
    return device_dict_list
"""


def parser_show_interface_description (output, template):
    import textfsm
    result = []
    f = open(template)
    re_table = textfsm.TextFSM(f)
    result_header = re_table.header
    result.append(result_header)
    result_output = re_table.ParseText(output)
    result.extend(result_output)
    return result


"""
def send_show_command(device_list, command):
    key=[]
    result_comm_dict={}
    for dict in device_list:
        key = list(dict.keys())
        print('Подключаемся к устройству с IP адресом {} '.format(dict[key[1]]))
        with ConnectHandler(**dict) as telnet:
            telnet.enable()
            result_comm = telnet.send_command(command)
            print(result_comm)
            result_comm_dict[dict[key[1]]]=result_comm
    #print(result_comm_dict)
    return(result_comm_dict)
"""


def send_show_command(IP, command, Username, Password, queue):
    output_list = []
    #for IP in address_list:
    #print('Connection to device {}'.format(IP))
    with telnetlib.Telnet(IP) as t:
        t.read_until(b'Username:')
        t.write(Username + b'\n')

        t.read_until(b'Password:')
        t.write(Password + b'\n')
        #t.write(b'enable\n')

        #t.read_until(b'Password:')
        #t.write(ENABLE_PASS + b'\n')
        t.write(b'terminal length 0\n')
        t.write(command + b'\n')

        time.sleep(5)

        output = t.read_very_eager().decode('utf-8')
        queue.put({IP: output})

def conn_threads(function, devices, command, Username, Password, limit=5):
    from queue import Queue

    devices_groups = (devices[idx:idx+limit]
                      for idx in range(0, len(devices), limit))

    for group in devices_groups:
        print(group)
        threads = []
        q = Queue()

        for device in devices:
            # Передаем очередь как аргумент, функции
            th = threading.Thread(target=function, args=(device, command, Username, Password, q))
            th.start()
            threads.append(th)

        for th in threads:
            th.join()

    results = []
    # Берем результаты из очереди и добавляем их в список results
    for t in threads:
        results.append(q.get())

    return results

listing_out = []
device_list = []
device = []
device_template = ['device_type', 'ip', 'username', 'password', 'enable']


print('Insert name for output file:')
output_file_name = input('output file name:')
print('Insert Username and Password')
Username = input('Username:').encode('utf-8')
Password = getpass.getpass().encode('utf-8')
print('Insert command for send to routers')
command = input('Command:').encode('utf-8')
template = 'Cisco_template_arp.txt'
file = sys.argv[1]
address_list = open_excel_routers(file)
"""
for address in address_list:
    device.append(address)
    device = device + Username.split() + Password.split() #+ enable_Password.split()
    device_list.append(device)
    #device = []
device_list = device_dict(device_list, device_template)
#command_listing = send_show_command(device_list, command)
"""

start_time = datetime.now()
output_command = conn_threads(send_show_command, address_list, command, Username, Password)
for dict in output_command:
    listing = parser_show_interface_description('\r\n'.join(list(dict.values())), template)
    listing_out.extend(listing)
output_file = save_output_to_excel(listing_out, output_file_name)

print('Скрипт выполнялся {}'.format(datetime.now() - start_time))

