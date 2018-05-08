import openpyxl
import sys
import getpass
import telnetlib
import time
import re
import threading
from datetime import datetime
from tabulate import tabulate
import clitable


def open_excel_routers(file):
    addresses = []
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    for row in range(1, sheet.max_row+1):
        addresses.append(sheet['A'+str(row)].value)
    return addresses

def save_output_to_excel(output_list, file_name):
    from openpyxl import Workbook
    wb = Workbook()
    sheet = wb.active
    for i in range(0, len(output_list)):
        for k in range(0, 6):
            sheet.cell(row=(i+1), column=(k+1)).value = output_list[i][k]
    wb.save(file_name)

def parser_show_interface_description (output, template):
    import textfsm
    result = []
    f = open(template)
    re_table = textfsm.TextFSM(f)
    result_header = re_table.header
    result.append(result_header)
    result_output = re_table.ParseText(output)
    result.extend(result_output)
    return result


def parser_show_interface_description_clitable (output, command):
    result = []
    cli_table = clitable.CliTable('index', 'templates')
    attributes = {'Command': command, 'Vendor': 'ZTE'}
    cli_table.ParseCmd(output, attributes)
    data_rows = [list(row) for row in cli_table]
    header = list(cli_table.header)
    result.append(header)
    result.extend(list(data_rows))
    return result


def conn_threads(function, devices, command, Username, Password, limit=5):
    from queue import Queue

    devices_groups = (devices[idx:idx+limit]
                      for idx in range(0, len(devices), limit))

    for group in devices_groups:
        threads = []
        q = Queue()

        for device in devices:
            # Передаем очередь как аргумент, функции
            th = threading.Thread(target=function, args=(device, command, Username, Password, q))
            th.start()
            threads.append(th)

        for th in threads:
            th.join()

    results = []
    # Берем результаты из очереди и добавляем их в список results
    for t in threads:
        results.append(q.get())

    return results


def send_show_command(IP, command, Username, Password, queue):
    output_list = []
    '''
    for IP in address_list:
    '''
    print('Connection to device {}'.format(IP))
    with telnetlib.Telnet(IP) as t:
        t.read_until(b'Username:')
        t.write(Username + b'\n')

        t.read_until(b'Password:')
        t.write(Password + b'\n')
        t.write(b'terminal length 0\n')
        t.write(command + b'\n')
        time.sleep(5)

        #output = t.read_very_eager()
        output = t.read_very_eager()
        queue.put({IP: output})

listing_out = []
device_list = []
device = []
device_template = ['device_type', 'ip', 'username', 'password', 'enable']

output_file_name = input('Output filename:')
print('Insert Username and Password')
Username = input('Username:').encode('utf-8')
Password = getpass.getpass().encode('utf-8')
print('Insert command for send to routers')
command = input('Command:').encode('utf-8')

template = 'ZTE_intf_descr_template.txt'
#file = sys.argv[1]
file='router_test_1.xlsx'
address_list = open_excel_routers(file)

start_time = datetime.now()
output_command = conn_threads(send_show_command, address_list, command, Username, Password)
print(output_command)


'''for dict in output_command:
    listing = parser_show_interface_description_clitable(re.sub('(\r\n {66})*', '', ''.join(list(dict.values()))), command.decode('utf-8'))
    listing_out.extend(listing)

print(listing_out)
output_file = save_output_to_excel(listing_out, output_file_name)

print('Скрипт выполнялся {}'.format(datetime.now() - start_time))
'''